'''  #																	||
---  #																	||
<(META)>:  #															||
	DOCid:   #						||
	name: Pheonix Molecules Controller Office Sheet Python Execution Document  #||
	description: >  #													||
		Automated generation of generic spreadsheets using configuration
		files and heavily defaulted options#	||

		Custom browsing tools for embeding in nchantrs and focused web
		interaction#	||

	expirary: <[expiration]>  #											||
	version: <[version]>  #												||
	path: <[LEXIvrs]>  #												||
	outline: <[outline]>  #												||
	authority: document|this  #											||
	security: sec|lvl2  #												||
	<(WT)>: -32  #														||
''' #																	||
# -*- coding: utf-8 -*-#												||
#===========================Core Modules================================||
import os, json
#===========================3rd Party Modules===========================||
import openpyxl as xl
from yaml import load as yload, dump as ydump#									||
try:#																			||
	from yaml import CLoader as Loader, CDumper as Dumper#						||
except:#																		||
	from yaml import Loader, Dumper#											||
#============================Pheonix Modules============================||
here = os.path.join(os.path.dirname(__file__),'')#								||
there = os.path.abspath(os.path.join('../../..'))#								||set path at pheonix level
version = '0.0.0.0.0.0'#														||
#=======================================================================||
l = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R',#	||
											'S','T','U','V','W','X','Y','Z']#	||
class book(object):
	def __init__(self):
		''' '''
def getLetter(coln):#															||
	cnt, ltr = -1, ''#															||
	while coln > 25:#															||
		coln = coln-26#															||
		cnt += 1#																||
#		if cnt > 25:#															||
		if cnt < 25:
			ltr += l[cnt]#														||
			cnt = 0#															||
	if cnt != -1:#																||
		ltr += l[cnt]#															||
	ltr += l[coln]#																||
	return ltr#																	||
class workBookWriter(object):
	def __init__(self):
		try:
			self.wb = xl.Workbook()
		except Exception as e:
			print('workbook create failed',e)
		self.createFormats()
	def creWS(self, name):
		''
#		print('Sheet Name',name)
#		if '/' in name:
#			name = name.replace('/','')
		name = sanSheetName(name)
		return self.wb.create_sheet(title=name)
	def createFormats(self, cfgs=None):
		''

		if cfgs == None:
			FORMATS = getYAML('{0}z-data_/xl.yaml'.format(here))
		else:
			FORMATS = getYAML('{0}'.format(cfgs['xl']))
#		cfg = 'z-data_/xl.yaml'.format(here)
#		FORMATS = config.instruct(cfg).load().dikt['areas']
		for style in sorted(FORMATS['areas'].keys()):
			cfg = FORMATS['areas'][style]
			self.wb.add_named_style(createStyle(style, cfg))
	def setColumnWidths(self, ws, mapp):
		''
		for col in sorted(mapp.keys()):
			ws.column_dimensions[col].width = mapp[col]
	def setRowHeights(self, ws, mapp):
		''
		for row in sorted(mapp.keys()):
			ws.row_dimensions[row].height = mapp[row]
	def mapTable2Cells(self, ws, datat):#								||
		''
		rcnt = 1#														||
		datat.sort()
		for row in datat:#												||
			ccnt = 1#													||
			row.sort()
			for col in row:#											||
				_ = ws.cell(column=ccnt, row=rcnt,)#	||
				setCellValue(_, col)
				ccnt += 1
			rcnt += 1
	def mapDict2Cells(self, ws, datad):
		''
		for key in sorted(datad.keys()):
#			print('mapDict2Cells Key',key)
			val = datad[key]
#			print('XLDict Key',key,'value',val)
#			print('mapDict2Cells',ws[key],ws[key].col_idx)
			try:
				_ = ws.cell(column=ws[key].col_idx, row=ws[key].row)#	||
			except Exception as e:
				print('Map Cell Failed key',key,e,ws[key])
			self.setCellValue(_, val)
	def setCellValue(self, cell, val, cfg=None):
		lock = 0
		if val == '' or val == None:
			lock = 2
		if '=' in str(val) and lock == 0:
			cell.set_explicit_value('{0}'.format(val), data_type='f')
			cell.number_format = '#0.00000'
			lock = 1
		try:
			if lock == 0 and str(val)[0].isdigit() or str(val)[1].isdigit():
				cell.set_explicit_value('{0}'.format(val), data_type='n')
				cell.number_format = '#0.00000'
				lock = 1
		except:
			pass
		if lock == 0 or lock == 2:
			cell.set_explicit_value('{0}'.format(val), data_type='s')
	def save(self, path=None,  name='QLFs.xlsx'):
		try:
			self.wb.remove(self.wb.get_sheet_by_name('Sheet'))
		except:
			pass
		if path == None:
			path = self.opath
		self.wb.save(filename='{0}/{1}'.format(path, name))#									||
def columnCollector(column, srow, erow):
	''
	vals = []
#	column.sort()
	for cval in column:
		if cval.row >= srow:
			vals.append(cval.value)
		if cval.row > erow:
			break
	return vals
def columnScanner(column, slimit, elimit, inc=False):
	''
	lock, startrow, endrow = 0, None, None
	for val in reversed(list(column)):
		if lock == 0:
			if val.value != elimit:
				lock = 1
				endrow = val.row
				continue
		if lock == 1:
			if val.value == slimit:
				startrow = val.row
	if endrow == 1:
		startrow = 1
	elif inc == False:
		if startrow is not None:
			startrow += 1
		if endrow is not None:
			endrow = endrow - 1
	return startrow, endrow
def creScatterChart(ws, reports, scol, ecol, srow, erow, minn, maxx, lsl, usl, brdrcol, cfg={}):
	''
	if cfg != {}:
		if cfg['how'] == 2:
			yval_minrow = srow
			yval_maxrow = erow-2
			ycol_offset = ecol-scol+4
			xval_minrow = srow+1
			xval_maxrow = erow-2
			xcol_offset = 4
			ecol = ecol-3
	else:
		yval_minrow = erow+1
		yval_maxrow = erow+erow-srow+1
		ycol_offset = 0
		xval_minrow = srow+1
		xval_maxrow = erow
		xcol_offset = 0
		ecol = ecol+2
	chart = xl.chart.ScatterChart()
	chart.style = 42
	for i in range(scol,ecol):
		values = xl.chart.Reference(ws, min_col=i+ycol_offset, min_row=yval_minrow, max_row=yval_maxrow)
		xvalues = xl.chart.Reference(ws, min_col=i+xcol_offset, min_row=xval_minrow, max_row=xval_maxrow)
		series = xl.chart.Series(values, xvalues=xvalues, title_from_data=True)
		chart.series.append(series)
	for s in chart.series:
		s.marker.symbol = "circle"
		s.marker.size = 7
		s.graphicalProperties.line.noFill = True
	lslv = xl.chart.Reference(ws, min_col=1, min_row=39, max_row=40)
	mintomax = xl.chart.Reference(ws, min_col=1, min_row=41, max_row=42)
	uslv = xl.chart.Reference(ws, min_col=brdrcol, min_row=39, max_row=40)
	series = xl.chart.Series(mintomax, xvalues=lslv, title_from_data=False)
	lineProp = xl.drawing.line.LineProperties(solidFill = xl.drawing.colors.ColorChoice(prstClr="yellow"))
	series.graphicalProperties.line = lineProp
	chart.series.append(series)
	series = xl.chart.Series(mintomax, xvalues=uslv, title_from_data=False)
	lineProp = xl.drawing.line.LineProperties(solidFill = xl.drawing.colors.ColorChoice(prstClr="red"))
	series.graphicalProperties.line = lineProp
	chart.series.append(series)
	chart.legend = None
	chart.y_axis.delete = True
	if minn > lsl:
		minn = Decimal(lsl)
	if maxx < usl:
		maxx = Decimal(usl)
	if not 'how' in cfg.keys():
		extra = (maxx+minn)/2*Decimal(0.0025)
		minclip = round(minn-extra, 5)
		maxclip = round(maxx+extra, 5)
		chart.x_axis.scaling.min = minclip
		chart.x_axis.scaling.max = maxclip
	chart.height = 10
	chart.width = 20
	chart.x_axis.crosses = 'min'
	chart.y_axis.crosses = 'min'
	props = xl.chart.shapes.GraphicalProperties(solidFill="404040")#"DAE3F3")
	chart.plot_area.graphicalProperties = props
	props = xl.chart.shapes.GraphicalProperties(solidFill='404040')#"767171")
	chart.graphical_properties = props
	return chart
def styleSheet(ws, stylemap, cfg=None):
	''
	ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
	ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
	ws.page_setup.fitToHeight = 0
	ws.page_setup.fitToWidth = 1
	if cfg != None:
		ws.sheet_properties.tabColor = cfg['tabColor']
	for area in sorted(stylemap.keys()):
		cfg = stylemap[area]
		for row in range(cfg['range']['minrow'], cfg['range']['maxrow']):
			for col in range(cfg['range']['mincol'], cfg['range']['maxcol']):
				cell = ws[getLetter(col)+str(row)]
#				print('Cell',cell)
				try:
					cell.style = cfg['style']
#					print('CELL', cell)
				except Exception as e:
#					print('CELL Tuple Problem', e,cell)
					cell[0].style = cfg['style']
				if 'datatype' in cfg.keys():
					cell.set_explicit_value(cell.value, data_type=cfg['datatype'])
		if 'rules' in cfg.keys():
			if cfg['rules'] != None:
				for rule in sorted(cfg['rules'].keys()):
					params = cfg['rules'][rule]
					brng = getLetter(cfg['range']['mincol'])+str(cfg['range']['minrow'])
					erng = getLetter(cfg['range']['maxcol'])+str(cfg['range']['maxrow'])
					rang = '$'+brng+':$'+erng
					if params['type'] == 'CellIsRule':
						colors = params['params']['fill']
						params['params']['fill'] = xl.styles.PatternFill(**colors)
						try:
							params['params']['font'] = xl.styles.Font(color=params['params']['font'])
						except:
							pass
						rule = xl.formatting.rule.CellIsRule(**params['params'])
					else:
						fill = xl.styles.PatternFill(bgColor=params['params']['color']['fill'],
													fgColor=params['params']['color']['fill'])
						font = xl.styles.Font(color=params['params']['color']['font'])
						dxf = xl.styles.differential.DifferentialStyle(font=font, fill=fill)
						rule = xl.formatting.rule.Rule(type=params['params']['type'], dxf=dxf)
						rule.formula = params['params']['formula']
					ws.conditional_formatting.add(rang,rule)
def createStyle(name, cfg):#											||
	cfg = cfg['Styles']
	style = xl.styles.NamedStyle(name=name)#							||
	style.font = xl.styles.Font(**cfg['Font'])#							||
	style.fill = xl.styles.PatternFill(**cfg['Fill'])#					||
	try:
		left = cfg['Border']['left']
		right = cfg['Border']['right']
		top = cfg['Border']['top']
		bottom = cfg['Border']['bottom']
	except:
		left, right, top, bottom = {}, {}, {}, {}
	style.border = xl.styles.Border(left=xl.styles.Side(**left),#					||
										right=xl.styles.Side(**right),
										top=xl.styles.Side(**top),#				||
										bottom=xl.styles.Side(**bottom),)#		||
	style.alignment = xl.styles.Alignment(**cfg['Alignment'])#			||
	style.protection = xl.styles.Protection(**cfg['Protection'])
	return style
def wsScanner(ws):
	records = []
	rows = ws.values
	rcnt = 1
	while True:
		try:
			if rcnt == 1:
				columns = [x for x in list(next(rows)) if x != None]
			else:
				records.append(list(next(rows))[:len(columns)])
			rcnt += 1
		except Exception as e:
			print(e)
			break
	return columns, records
def creWSTable(data, sheetTMPLT=None):
	styles = sheetTMPLT['styles']
	fdmap, sizemaps = {}, {'columnmap': {}, 'rowmap': {}}
	mapped, coln, rown = mapAREA(sheetTMPLT['map']['header'], 0, 0, 'tree')
	fdmap.update(mapped)#												||
#	print('Data', data)
	mapped, coln, rown = mapAREA(data, 0, rown, 'table')
	fdmap.update(mapped)
	stmap = {'background': styles['background'], 'header': styles['header'], 'data': styles['data']}#	||
	stmap['background']['range']['minrow'] = 1
	stmap['background']['range']['maxrow'] = rown + 5
	stmap['background']['range']['mincol'] = 0
	stmap['background']['range']['maxcol'] = coln + 5
	stmap['header']['range']['minrow'] = 1
	stmap['header']['range']['mincol'] = 0
	stmap['header']['range']['maxrow'] = 2
	stmap['header']['range']['maxcol'] = coln
	stmap['data']['range']['minrow'] = 2
	stmap['data']['range']['mincol'] = 0
	stmap['data']['range']['maxrow'] = rown
	stmap['data']['range']['maxcol'] = coln
	sizes = sheetTMPLT['size']
	sizemaps = {'columnmap': sizes['cols'], 'rowmap': {}}
	for i in range(rown):
		if i not in sizemaps['rowmap'].keys():#							||
			sizemaps['rowmap'][i] = 15.75#								||
	return fdmap, stmap, sizemaps
def mapAREA(data, scol=0, srow=1, how='table'):
	''
	mapp = {}
	if how == 'table':
		rown = srow
#		data.sort()
		for rowd in data:
			coln = scol
			rowd = [str(x) for x in rowd]
#			rowd.sort()
			for colv in rowd:
				position = getLetter(coln)+str(rown)
				mapp[position] = colv
				coln += 1
			rown += 1
	elif how == 'tree':
		if not isinstance(data, dict):
#			print('Tree Data', type(data))
			data = json.loads(data)
		for rowd in sorted(data.keys()):
			rowv = data[rowd]
			for cold in sorted(rowv.keys()):
				colv = rowv[cold]
				coln = scol+int(cold)
				rown = srow+int(rowd)
				position = getLetter(coln)+str(rown)
				mapp[position] = colv
		rown += 1
	return mapp, coln, rown
def mapTableAREA(name, head=None, body=None, foot=None, aCNT=0,#		||
								scol=0, rown=1, fdmap={}, cfgs=None):#	||
	'Create a cell dictionary mapped to a column/row grid'#				||
	if cfgs == None:
		cfg = getYAML('{0}z-data_/qlf.yaml'.format(here))
	else:
		cfg = getYAML('{0}'.format(cfgs['qlf']))
	fdmap, maxcoln = {}, 0
	styles = cfg['TMPLTs']['sectns'][name]['styles']#					||
	hdrAREA = 'header{0}'.format(aCNT)#									||
	bdyAREA = 'body{0}'.format(aCNT)#									||
	ftrAREA = 'footer{0}'.format(aCNT)#									||
	stmap = {hdrAREA: styles['header'].copy(),
									bdyAREA: styles['data'].copy(),#	||
									ftrAREA: styles['header'].copy()}#	||
	if head != None:#													||
		stmap[hdrAREA]['range']['minrow'] = rown#						||
		stmap[hdrAREA]['range']['mincol'] = scol#						||
		headMPD, coln, rown = mapAREA(head, scol, rown, 'tree')#		||
		if coln > maxcoln:
			maxcoln = coln
		fdmap.update(headMPD)#											||
		stmap[hdrAREA]['range']['maxrow'] = rown#						||
	else:
		print('Head was None for',name)
	if body != None:#													||
		stmap[bdyAREA]['range']['minrow'] = rown#						||
		stmap[bdyAREA]['range']['mincol'] = scol#						||
		bodyMPD, bcoln, rown = mapAREA(body, scol, rown, 'table')#		||
		if coln > maxcoln:
			maxcoln = coln
		fdmap.update(bodyMPD)#											||
		stmap[bdyAREA]['range']['maxrow'] = rown#						||
	else:
		print('Body was None for',name)
	if foot != None:#													||
		stmap[ftrAREA]['range']['minrow'] = rown#						||
		stmap[ftrAREA]['range']['mincol'] = scol#						||
		footMPD, coln, rown = mapAREA(foot, scol, rown, 'tree')#		||
		if coln > maxcoln:
			maxcoln = coln
		fdmap.update(footMPD)#											||
		stmap[ftrAREA]['range']['maxrow'] = rown#						||
	else:
		print('Footer Was None for',name)
	stmap[hdrAREA]['range']['maxcol'] = maxcoln+1#						||
	stmap[bdyAREA]['range']['maxcol'] = maxcoln+1#						||
	stmap[ftrAREA]['range']['maxcol'] = maxcoln+1#						||
	mapp = {'value': fdmap, 'stmap': stmap}#							||
	return mapp, coln, rown, aCNT#										||
def sanSheetName(n):#													||
	''#													||
	subs = ['[', ']', '/']#												||
	for sub in subs:#													||
		if sub in n:#													||
			n = n.replace(sub, '')#										||
	return n#															||
def writeWS(name, wb, fdmap, stylemap=None, sizemaps=None):
	''
	sheet = wb.creWS(name)#														||
	if stylemap != None:
		styleSheet(sheet, stylemap)#											||
	if sizemaps != None:
		if 'columnmap' in sizemaps.keys():
			wb.setColumnWidths(sheet, sizemaps['columnmap'])
		if 'rowmap' in sizemaps.keys():
			wb.setRowHeights(sheet, sizemaps['rowmap'])
	wb.mapDict2Cells(sheet, fdmap)
def writeWB(data, path=None, name=None):
	''
	wb = workBookWriter()
	fdmap, coln, rown = mapAREA(data)
#	print('FDmap', fdmap.keys())
	writeWS('Data', wb, fdmap)
	wb.save(path, '{0}.xlsx'.format(name))
#	print('Path', path)
def getYAML(path):
	with open(path, 'r') as doc:#										||
		dikt = yload(doc.read().replace('\t','  '), Loader=Loader)#		||
	return dikt
